"""Crawler for chinamoney.com.cn FX implied rate curve data.

Uses Playwright + MS Edge to automate the browser-based Excel export:
  1. Launch Edge, navigate to the curve page
  2. Set date range and parameter dropdowns
  3. Click "导出Excel" and capture the downloaded .xlsx
  4. Parse with openpyxl, filter to target parameters, upsert to DB

Selector IDs verified against live page DOM on 2026-06-24.
"""

from __future__ import annotations

import logging
import tempfile
import traceback
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.curve import FxImpliedRate

from . import _playwright_helpers as pw

logger = logging.getLogger("optrade.crawler.china_money")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Dropdown values used on the page form.
# The export returns data for ALL parameter combinations; we filter
# to this specific set during parsing.
_TARGET_PARAMS = {
    "rmb_rate": "FR007利率互换收盘曲线",   # 人民币利率
    "spot": "即期询价报价均值",             # 即期汇率
    "swap": "掉期点",                      # 远/掉期点
}

# Number of past days to seed when the database table is empty.
_INITIAL_SEED_DAYS = 30

# ---------------------------------------------------------------------------
# XLSX column map (column index → internal dict key)
#
# Verified against actual export (2026-06-24) — exactly 10 columns:
#   0: 日期              1: 货币对
#   2: 人民币利率         3: 即期汇率
#   4: 远/掉期点          5: 期限
#   6: 隐含利率(%)        7: 人民币利率(%)
#   8: 即期汇率 (numeric)  9: 远/掉期点(Pips)
# ---------------------------------------------------------------------------

_COLUMN_MAP: dict[int, str] = {
    0: "curve_date",
    1: "ccy_pair",
    2: "rmb_rate_label",
    3: "spot_label",
    4: "swap_label",
    5: "tenor",
    6: "foreign_implied_rate",
    7: "cny_risk_free_rate",
    8: "spot_rate",
    9: "swap_points",
}

# Fields that should be converted to float (column indices).
_NUMERIC_COLS: set[int] = {6, 7, 8, 9}

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class ChinaMoneyCrawler:
    """Orchestrate Playwright-based crawling of chinamoney FX implied rates."""

    def __init__(self, headless: bool = True) -> None:
        self._headless = headless

    # ---- date helpers --------------------------------------------------

    @staticmethod
    def determine_missing_dates(session: Session) -> list[date]:
        """Return every calendar date from (max stored + 1) through today.

        When the table is empty the last *_INITIAL_SEED_DAYS* days are
        returned so the initial fill catches up recent history.
        """
        result = session.exec(
            select(FxImpliedRate.curve_date).order_by(
                FxImpliedRate.curve_date.desc(),
            )
        ).first()

        today = date.today()
        if result:
            start = result + timedelta(days=1)
        else:
            start = today - timedelta(days=_INITIAL_SEED_DAYS)

        if start > today:
            return []

        days: list[date] = []
        cursor = start
        while cursor <= today:
            days.append(cursor)
            cursor += timedelta(days=1)
        return days

    # ---- main entry point ----------------------------------------------

    async def crawl_all_missing(self, session: Session) -> dict:
        """Detect gaps, crawl chinamoney, parse & upsert.

        Returns a dict compatible with ``CrawlResult``.
        """
        missing = self.determine_missing_dates(session)
        if not missing:
            logger.info("No missing dates — crawl skipped")
            return {
                "status": "success",
                "dates_fetched": [],
                "dates_skipped": [],
                "records_added": 0,
                "error_message": None,
            }

        logger.info(
            "Missing date range: %s → %s (%d days)",
            missing[0],
            missing[-1],
            len(missing),
        )

        try:
            xlsx_bytes = await self._download_xlsx_via_browser(
                missing[0], missing[-1],
            )
        except Exception as exc:
            logger.error(
                "Browser download failed: %s\n%s",
                exc,
                traceback.format_exc(),
            )
            return {
                "status": "error",
                "dates_fetched": [],
                "dates_skipped": missing,
                "records_added": 0,
                "error_message": str(exc),
            }

        try:
            records = self.parse_xlsx(xlsx_bytes)
        except Exception as exc:
            logger.error(
                "XLSX parse failed: %s\n%s",
                exc,
                traceback.format_exc(),
            )
            return {
                "status": "error",
                "dates_fetched": missing,
                "dates_skipped": [],
                "records_added": 0,
                "error_message": f"Parse error: {exc}",
            }

        added = self._upsert_rates(session, records)
        logger.info(
            "Crawl complete: %d records added (out of %d parsed)",
            added,
            len(records),
        )
        return {
            "status": "success",
            "dates_fetched": missing,
            "dates_skipped": [],
            "records_added": added,
            "error_message": None,
        }

    # ---- browser automation --------------------------------------------

    async def _download_xlsx_via_browser(
        self, date_from: date, date_to: date,
    ) -> bytes:
        """Automate the chinamoney Excel export with Playwright + Edge."""
        pw_browser, pw_instance = await pw.launch_browser(headless=self._headless)

        try:
            page = await pw_browser.new_page()
            await pw.navigate_to_page(page)
            await pw.wait_for_page_ready(page)

            # Fill form.
            await pw.set_date_range(page, date_from, date_to)
            await pw.set_dropdowns(
                page,
                currency="全部",
                rmb_rate=_TARGET_PARAMS["rmb_rate"],
                spot=_TARGET_PARAMS["spot"],
                swap=_TARGET_PARAMS["swap"],
            )

            # Click export and capture the download.
            async with page.expect_download(timeout=120_000) as download_info:
                await pw.click_export_button(page)

            download = await download_info.value
            logger.info("Download captured: %s", download.suggested_filename)

            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False,
            ) as tmp:
                tmp_path = Path(tmp.name)
            try:
                await download.save_as(str(tmp_path))
                data = tmp_path.read_bytes()
            finally:
                tmp_path.unlink(missing_ok=True)

            logger.info("XLSX size: %d bytes", len(data))
            return data

        finally:
            await pw_browser.close()
            await pw_instance.stop()

    # ---- XLSX parsing --------------------------------------------------

    @staticmethod
    def parse_xlsx(xlsx_bytes: bytes) -> list[dict]:
        """Parse an XLSX byte stream into field dicts, filtered to target params.

        The chinamoney export includes ALL parameter combinations in one
        file.  We keep only rows matching *_TARGET_PARAMS* (FR007 + 即期询价
        报价均值 + 掉期点).

        Footer rows (数据来源, www.chinamoney.com.cn) are skipped.
        """
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False,
        ) as tmp:
            tmp.write(xlsx_bytes)
            tmp_path = Path(tmp.name)

        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            tmp_path.unlink(missing_ok=True)

        if len(rows) < 2:
            logger.warning("XLSX has fewer than 2 rows – no data to parse")
            return []

        # Row 0 is the header; rows 1.. are data or footer.
        records: list[dict] = []
        for row in rows[1:]:
            record = ChinaMoneyCrawler._parse_data_row(row)
            if record is None:
                continue

            # Filter to target parameter combination.
            if (
                record.get("rmb_rate_label") != _TARGET_PARAMS["rmb_rate"]
                or record.get("spot_label") != _TARGET_PARAMS["spot"]
                or record.get("swap_label") != _TARGET_PARAMS["swap"]
            ):
                continue

            # Clean up: remove label fields, keep only model columns.
            record.pop("rmb_rate_label", None)
            record.pop("spot_label", None)
            record.pop("swap_label", None)

            # Extract foreign_currency from ccy_pair.
            ccy_pair: str | None = record.pop("ccy_pair", None)
            if ccy_pair:
                record["foreign_currency"] = _extract_foreign_currency(ccy_pair)

            if record.get("curve_date") and record.get("foreign_currency") and record.get("tenor"):
                records.append(record)

        logger.debug(
            "Parsed %d records (filtered from %d data rows)",
            len(records),
            len(rows) - 1,
        )
        return records

    @staticmethod
    def _parse_data_row(row: tuple) -> dict | None:
        """Convert one XLSX row into a dict using the fixed column map.

        Returns None for footer rows or rows with unparseable dates.
        """
        record: dict = {}

        for idx, value in enumerate(row):
            field = _COLUMN_MAP.get(idx)
            if field is None:
                continue

            if value is None:
                record[field] = None
            elif isinstance(value, datetime):
                record[field] = value.date()
            elif isinstance(value, str):
                s = value.strip()
                if not s:
                    record[field] = None
                elif idx in _NUMERIC_COLS:
                    try:
                        record[field] = float(s)
                    except ValueError:
                        record[field] = None
                else:
                    record[field] = s
            elif idx in _NUMERIC_COLS and isinstance(value, (int, float)):
                record[field] = float(value)
            else:
                record[field] = value

        # ---- footer guard ----
        raw_date = record.get("curve_date")
        if raw_date is None or isinstance(raw_date, str) and not _looks_like_date(raw_date):
            return None

        # ---- parse date ----
        if isinstance(raw_date, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
                try:
                    record["curve_date"] = datetime.strptime(
                        raw_date.strip(), fmt,
                    ).date()
                    break
                except ValueError:
                    continue
            else:
                logger.debug("Could not parse date string: %r", raw_date)
                return None

        return record

    # ---- DB persistence ------------------------------------------------

    @staticmethod
    def _upsert_rates(session: Session, records: list[dict]) -> int:
        """Insert new rates, skipping existing (curve_date, currency, tenor)."""
        added = 0
        for rec in records:
            if not all(k in rec for k in ("curve_date", "foreign_currency", "tenor")):
                continue

            existing = session.exec(
                select(FxImpliedRate).where(
                    FxImpliedRate.curve_date == rec["curve_date"],
                    FxImpliedRate.foreign_currency == rec["foreign_currency"],
                    FxImpliedRate.tenor == rec["tenor"],
                )
            ).first()
            if existing is not None:
                continue

            rate = FxImpliedRate(
                curve_date=rec["curve_date"],
                foreign_currency=rec["foreign_currency"],
                tenor=rec["tenor"],
                foreign_implied_rate=rec.get("foreign_implied_rate"),
                cny_risk_free_rate=rec.get("cny_risk_free_rate"),
                spot_rate=rec.get("spot_rate"),
                swap_points=rec.get("swap_points"),
                source="chinamoney",
            )
            session.add(rate)
            added += 1

        if added:
            session.commit()
        return added


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _extract_foreign_currency(ccy_pair: str) -> str:
    """``"USD.CNY"`` → ``"USD"``, ``"EUR/CNY"`` → ``"EUR"``."""
    for sep in (".", "/"):
        if sep in ccy_pair:
            return ccy_pair.split(sep)[0].strip().upper()
    return ccy_pair.strip().upper()


def _looks_like_date(s: str) -> bool:
    """Return True if *s* starts with a year-like pattern (e.g. ``2026-…``)."""
    return len(s) >= 8 and s[:4].isdigit()
