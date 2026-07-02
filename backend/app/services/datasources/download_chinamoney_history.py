"""Historical chinamoney FX implied rate downloader.

Chinamoney restricts each export to roughly one month of data.  This script
splits a long date range into monthly chunks, downloads each chunk via the
existing Playwright automation, then merges all workbooks into a single .xlsx.

Example:
    python -m app.services.datasources.download_chinamoney_history \
        --start 2025-01-01 --end 2026-07-01 --output merged.xlsx \
        --save-dir ./chinamoney_chunks
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import tempfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.datasources.china_money_crawler import (
    _TARGET_PARAMS,
    ChinaMoneyCrawler,
)
from app.services.datasources import _playwright_helpers as pw

logger = logging.getLogger("optrade.crawler.history")

# Maximum window size for a single chinamoney export.  Keep it a little under
# a full calendar month to avoid edge-case rejections.
_CHUNK_DAYS = 28


async def _download_one_chunk(
    page,
    date_from: date,
    date_to: date,
    save_dir: Path | None = None,
) -> tuple[bytes, Path | None]:
    """Download a single chinamoney Excel chunk for the given date range.

    Returns the raw file bytes and, if *save_dir* is provided, the path of
    the saved original file.
    """
    await pw.set_date_range(page, date_from, date_to)
    await pw.set_dropdowns(
        page,
        currency="全部",
        rmb_rate=_TARGET_PARAMS["rmb_rate"],
        spot=_TARGET_PARAMS["spot"],
        swap=_TARGET_PARAMS["swap"],
    )
    await pw.click_query_button(page)

    async with page.expect_download(timeout=120_000) as download_info:
        await pw.click_export_button(page)

    download = await download_info.value
    logger.info("Chunk %s → %s downloaded: %s", date_from, date_to, download.suggested_filename)

    # Save original file if requested.
    saved_path: Path | None = None
    if save_dir is not None:
        save_dir.mkdir(parents=True, exist_ok=True)
        saved_path = save_dir / f"{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        await download.save_as(str(saved_path))
        logger.info("Original chunk saved to %s", saved_path)
        return saved_path.read_bytes(), saved_path

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        await download.save_as(str(tmp_path))
        return tmp_path.read_bytes(), None
    finally:
        tmp_path.unlink(missing_ok=True)


async def download_history(
    start_date: date,
    end_date: date,
    headless: bool = True,
    save_dir: Path | None = None,
) -> list[bytes]:
    """Download chinamoney data in monthly chunks and return raw .xlsx bytes."""
    chunks: list[tuple[date, date]] = []
    cursor = start_date
    while cursor <= end_date:
        chunk_end = min(cursor + timedelta(days=_CHUNK_DAYS - 1), end_date)
        chunks.append((cursor, chunk_end))
        cursor = chunk_end + timedelta(days=1)

    logger.info(
        "Splitting %s → %s into %d chunk(s) of max %d days",
        start_date, end_date, len(chunks), _CHUNK_DAYS,
    )

    pw_browser, pw_instance = await pw.launch_browser(headless=headless)
    page = await pw_browser.new_page()
    await pw.navigate_to_page(page)
    await pw.wait_for_page_ready(page)

    results: list[bytes] = []
    try:
        for idx, (df, dt) in enumerate(chunks, start=1):
            logger.info("Downloading chunk %d/%d: %s → %s", idx, len(chunks), df, dt)
            data, _ = await _download_one_chunk(page, df, dt, save_dir=save_dir)
            results.append(data)
    finally:
        await pw_browser.close()
        await pw_instance.stop()

    return results


def merge_xlsx_files(xlsx_bytes_list: list[bytes], output_path: Path) -> None:
    """Merge several chinamoney .xlsx byte streams into one workbook.

    The first file supplies the header row.  All subsequent files have their
    header row skipped to avoid duplication.
    """
    if not xlsx_bytes_list:
        raise ValueError("No .xlsx data to merge")

    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "历史数据"
    header_written = False

    for idx, raw in enumerate(xlsx_bytes_list):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = Path(tmp.name)

        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                logger.warning("Chunk %d is empty — skipping", idx + 1)
                continue

            start_row = 0 if not header_written else 1
            for row in rows[start_row:]:
                merged_ws.append(list(row))
            header_written = True
        finally:
            tmp_path.unlink(missing_ok=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged_wb.save(output_path)
    logger.info("Merged workbook saved to %s (%d sheets, %d rows)",
                output_path, len(merged_wb.sheetnames), merged_ws.max_row)


def _month_chunks(start_date: date, end_date: date) -> list[tuple[date, date]]:
    """Split range into calendar-month chunks.

    chinamoney accepts roughly one month at a time; using calendar months
    makes the chunks easy to reason about.
    """
    chunks: list[tuple[date, date]] = []
    cursor = date(start_date.year, start_date.month, 1)
    while cursor <= end_date:
        if cursor.month == 12:
            next_month = date(cursor.year + 1, 1, 1)
        else:
            next_month = date(cursor.year, cursor.month + 1, 1)
        month_end = next_month - timedelta(days=1)
        chunk_start = max(cursor, start_date)
        chunk_end = min(month_end, end_date)
        chunks.append((chunk_start, chunk_end))
        cursor = next_month
    return chunks


async def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download historical chinamoney FX implied rate curves and merge into one Excel.",
    )
    parser.add_argument("--start", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--output", default="chinamoney_history.xlsx", help="Output .xlsx path")
    parser.add_argument("--visible", action="store_true", help="Run browser visibly (not headless)")
    parser.add_argument("--save-dir", help="Directory to keep original downloaded .xlsx chunks")
    parser.add_argument("--use-existing", help="Directory of already-downloaded .xlsx files to merge")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

    start_date = date.fromisoformat(args.start)
    end_date = date.fromisoformat(args.end)
    output_path = Path(args.output)

    save_dir = Path(args.save_dir) if args.save_dir else None

    if args.use_existing:
        files = sorted(Path(args.use_existing).glob("*.xlsx"))
        xlsx_list = [f.read_bytes() for f in files]
        logger.info("Loaded %d existing file(s) from %s", len(xlsx_list), args.use_existing)
    else:
        xlsx_list = await download_history(
            start_date, end_date, headless=not args.visible, save_dir=save_dir,
        )

    merge_xlsx_files(xlsx_list, output_path)
    print(f"Saved merged file: {output_path.absolute()}")


if __name__ == "__main__":
    asyncio.run(main())
